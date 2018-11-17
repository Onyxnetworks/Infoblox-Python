import openpyxl
from openpyxl import Workbook
import requests, json, time
from getpass import getpass
import argparse

# Ignore SSL Errors
requests.packages.urllib3.disable_warnings()

class colour:
   PURPLE = '\033[95m'
   CYAN = '\033[96m'
   DARKCYAN = '\033[36m'
   BLUE = '\033[94m'
   GREEN = '\033[92m'
   YELLOW = '\033[93m'
   RED = '\033[91m'
   BOLD = '\033[1m'
   UNDERLINE = '\033[4m'
   END = '\033[0m'

def ERRORS():
    print(colour.RED + 'Errors found. Exiting script.' + colour.END)
    time.sleep(2)
    exit()

def BANNER():
    print'>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>'
    print 'Cisco ACI'
    print 'Tenant Fault Finder '
    print'>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>'
    time.sleep(1)

def USERNAME_PASSWORD():
    print '\nPlease enter adm credentials.'
    ADM_USER = raw_input("\nEnter Username: ")
    ADM_PASS = getpass("\nEnter Password: ")

    return ADM_USER, ADM_PASS	

def APIC_LOGIN(BASE_URL, APIC_USERNAME, APIC_PASSWORD, HEADERS):
    # Log into APIC and generate a cookie for future requests
    login_url = BASE_URL + 'aaaLogin.json'
    auth = {"aaaUser": {"attributes": {"name": APIC_USERNAME, "pwd": APIC_PASSWORD}}}
    auth_payload = json.dumps(auth)
    try:
        post_response = requests.post(login_url, data=auth_payload, headers=HEADERS, verify=False)
        # Take token from response to use for future authentications
        payload_response = json.loads(post_response.text)
        token = payload_response['imdata'][0]['aaaLogin']['attributes']['token']
        APIC_COOKIE = {}
        APIC_COOKIE['APIC-Cookie'] = token
        print('Connected Successfully.')

    except:
        exit(colour.RED + 'Unable to connect to APIC. Please check your credentials' + colour.END)

    return APIC_COOKIE

def GET_FAULTS(BASE_URL, APIC_COOKIE, TENANT_NAME, HEADERS):
    GET_FAULTS_BASE_URL = BASE_URL + 'node/mo/uni/tn-{0}.json?query-target=subtree&rsp-subtree-include=faults,no-scoped'.format(TENANT_NAME)

    try:
        get_response = requests.get(GET_FAULTS_BASE_URL, cookies=APIC_COOKIE, headers=HEADERS, verify=False)
        GET_FAULTS_RESPONSE = json.loads(get_response.text)
        return GET_FAULTS_RESPONSE


    except:
        exit(colour.RED + 'Failed to get faults for tenant: ' + TENANT_NAME + colour.END)	



def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("-v", "-verbose", help="Increase output verbosity", action="store_true")
    args = parser.parse_args()
    if args.v:
	    print('\nDebugging Enabled\n')
    ERROR = False
    HEADERS = {'content-type': 'application/json'}
    SANDBOX = 'https://sandboxapicdc.cisco.com/api/'
    DC = raw_input('DC: (DC1/DC2) ')
    TENANT_LIST = raw_input('Enter Tenant seperated by a comma: ').split(',')
    USER_DETAILS = USERNAME_PASSWORD()
    APIC_USERNAME = USER_DETAILS[0]
    APIC_PASSWORD = USER_DETAILS[1]

    # Select DC to to run script against
    if DC.upper() == 'DC1':
        BASE_URL = DC1
    elif DC.upper() == 'DC2':
        BASE_URL = DC2
    elif DC.upper() == 'SANDBOX':
        BASE_URL = SANDBOX
    elif DC.upper() == 'LAB':
        BASE_URL = LAB
    else:
        print('\nUnknown location selected. please choose from the following:')
        print('DC1 | DC2 | LAB | SANDBOX')
        exit()
	
    APIC_COOKIE = APIC_LOGIN(BASE_URL, APIC_USERNAME, APIC_PASSWORD, HEADERS)
    if APIC_COOKIE:
        time.sleep(2)
        print('\nSuccessfully generated authentication cookie')
    else:
        exit(colour.RED + '\nUnable to connect to APIC. Please check your credentials' + colour.END)
	
    wb = Workbook()
    print(colour.BOLD + '\nLooping through Tenants and building workbook.\n' + colour.END)

    for TENANT in TENANT_LIST:
        r = 3
        TENANT_NAME = TENANT
        GET_FAULTS_RESPONSE = GET_FAULTS(BASE_URL, APIC_COOKIE, TENANT_NAME, HEADERS)
        ws = wb.create_sheet(title=TENANT)
        print('Tenant: '+ TENANT_NAME + ' added as excel sheet.')
        print('Total Errors: ' +  GET_FAULTS_RESPONSE['totalCount'])
        ws['A2'] = 'Cause'
        ws['B2'] = 'Severity'
        ws['C2'] = 'Creation Time'
        ws['D2'] = 'Description'
        ws['E2'] = 'Code'
        ws['F2'] = 'Object'
        ws['A1'] = 'Total Faults: '
        ws['B1'] = GET_FAULTS_RESPONSE['totalCount']
        for FAULT in GET_FAULTS_RESPONSE['imdata']:
			if FAULT.keys() == [u'faultDelegate']:
				ws.cell(row=r, column=1).value = FAULT['faultDelegate']['attributes']['cause']
				ws.cell(row=r, column=3).value = FAULT['faultDelegate']['attributes']['created']
				ws.cell(row=r, column=6).value = FAULT['faultDelegate']['attributes']['dn']
				ws.cell(row=r, column=4).value = FAULT['faultDelegate']['attributes']['descr']
				ws.cell(row=r, column=5).value = FAULT['faultDelegate']['attributes']['code']
				ws.cell(row=r, column=2).value = FAULT['faultDelegate']['attributes']['severity']
			
			elif FAULT.keys() == [u'faultInst']:
 				ws.cell(row=r, column=1).value = FAULT['faultInst']['attributes']['cause']
				ws.cell(row=r, column=3).value = FAULT['faultInst']['attributes']['created']
				ws.cell(row=r, column=6).value = FAULT['faultInst']['attributes']['dn']
				ws.cell(row=r, column=4).value = FAULT['faultInst']['attributes']['descr']
				ws.cell(row=r, column=5).value = FAULT['faultInst']['attributes']['code']
				ws.cell(row=r, column=2).value = FAULT['faultInst']['attributes']['severity']        
			
			else:
				print(colour.YELLOW + '\nUnknown Key value: ' + FAULT.keys() + colour.END)
				pass

			r += 1
			
    # Remove empty sheet
    wb.remove(wb.get_sheet_by_name('Sheet'))

    # Save file
    try:
        SAVE_NAME = raw_input('Save As:  ').upper()
        print(colour.BOLD + 'Saving to file named "' + SAVE_NAME + '"' + colour.END)
        time.sleep(2)
        wb.save(SAVE_NAME + '.xlsx')
        time.sleep(2)
        print(colour.BOLD + 'File saved successfully' + colour.END)
		
    except:
        exit(colour.RED + '\nFailed to save file.' + colour.END)


if __name__ == "__main__":
    main()
